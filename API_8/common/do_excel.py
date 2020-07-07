# -*- coding: utf-8 -*-
# @Time    : 2020/6/18 13:59
# @Author  : zc
# @Email   : zc_9211@163.com
# @File    : do_excel.py
# @Software: PyCharm

from openpyxl import load_workbook
from API_8.common.http_request import HttpRequest


class Case:
    """
    测试用例类，每个测试用例，实际上就是它的一个实例
    """

    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None


class DoExcel:

    def __init__(self, file_name, sheet_name):
        # 异常处理
        try:
            self.file_name = file_name
            self.sheet_name = sheet_name
            self.wb = load_workbook(file_name)
            self.sheet = self.wb[sheet_name]
        except Exception as e:
            print('文件或sheet页选择错误，错误信息为{}'.format(e))

    def get_case(self):
        max_row = self.sheet.max_row
        cases = []  # 列表，存放所有的测试用例
        for i in range(2, max_row + 1):
            case = Case()  # case实例
            case.case_id = self.sheet.cell(i, 1).value
            case.title = self.sheet.cell(i, 2).value
            case.url = self.sheet.cell(i, 3).value
            case.data = self.sheet.cell(i, 4).value
            case.method = self.sheet.cell(i, 5).value
            case.expected = self.sheet.cell(i, 6).value
            cases.append(case)
        self.wb.close()
        return cases  # 返回case列表

    def write_result(self, row, actual, result):
        sheet = self.wb[self.sheet_name]
        sheet.cell(row, 7).value = actual
        sheet.cell(row, 8).value = result
        self.wb.save(filename=self.file_name)
        self.wb.close()

    # def close(self):
    #     self.wb.close()


if __name__ == '__main__':
    from API_3.common import contans

    do_excel = DoExcel(contans.case_file, 'login')
    cases = do_excel.get_case()
    http_request = HttpRequest()
    for case in cases:
        print(case.case_id)
        print(case.method)
        print(case.url)
        print(case.data)
        resp = http_request.request(case.method, case.url, data=case.data)
        print(resp.text)
        # print(type(resp.text))

        resp_dict = resp.json()  #返回字典

        actual = resp.text
        if case.expected == eval(actual)["message"]:
            do_excel.write_result(case.case_id + 1, actual, 'pass')
        else:
            do_excel.write_result(case.case_id + 1, actual, 'fail')

    # reslut = do_excel.write_result(2,3,4)
